import openpyxl  # used for Excel
import requests  # used to request info from site
from bs4 import BeautifulSoup  # used to parse the XML file from the Search API

# Open and load workbook/worksheet

workbookName = 'template.xlsx'  # Insert your Excel file name here.
workbook = openpyxl.load_workbook(workbookName)

worksheet = "Sheet"
worksheet = workbook[worksheet]

row = 2  # Skips title row

response = 'y'
while response == 'y':

    # requests info from OCLC WorldCat Search API
    book_ISBN = input("Book ISBN?: ")
    website_url = requests.get('http://classify.oclc.org/classify2/Classify?isbn=' + book_ISBN + '&summary=true').text
    soup = BeautifulSoup(website_url, 'lxml')

    # Find elements in XML file and output the attribute value within <work> tag
    for element in soup.find_all("classify"):
        for stat in element.find_all("work"):
            title = stat['title']
            print("Title: ", title)

            author = stat['author']
            print("Author: ", author)

            edition = stat['editions']
            print("Edition: ", edition)

    # insert information from API to Excel
    worksheet.cell(row=row, column=2).value = title
    worksheet.cell(row=row, column=3).value = author
    worksheet.cell(row=row, column=4).value = edition
    worksheet.cell(row=row, column=5).value = book_ISBN

    # If the user continues, move input to next row
    response = input("\nContinue? (Y/N): ")
    if response == "N":
        break 
    row += 1

# Save and close the Excel workbook
workbook.save(workbookName)
workbook.close

# TEST ISBN's
# 978-0-205-25949-6: Art book
# 978-1-2851990-2-3: Chem book
# 1782197567: Random Tom Hardy book